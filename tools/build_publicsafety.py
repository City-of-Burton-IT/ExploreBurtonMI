# Build public/info-publicsafety.json for the Public Safety (Fire & Rescue) dashboard.
#
# Source: City of Burton Fire Department records system -- Emergency Networking
# (Tyler) RMS. The FD runs two canned "standard reports" and exports them to
# .xlsx; this script transforms those exports into the committed dashboard JSON.
#   #1390  Incident Count and Percentage per Incident Type  (run for one full year)
#   #22    Incidents per Year                                (run for a multi-year range)
#
# GOVERNANCE (ADR-0001 + the public-safety discipline):
#   * Aggregates ONLY. The #1390 export is row-per-incident and carries a
#     "Location" column with street addresses -- this script NEVER reads it and
#     a guard fails loud if any address-like string reaches the output.
#   * The .xlsx exports are NOT committed (they contain incident addresses). Keep
#     them outside the repo (default: your Downloads folder).
#   * The output is written with "draft": true and the panel is held out of the
#     live deploy until the Fire Department signs off on the figures.
#
# Re-runnable (the site reads the committed JSON, never the RMS):
#   python tools/build_publicsafety.py
#   python tools/build_publicsafety.py --type-report PATH --year-report PATH
#
# Requires: openpyxl (see tools/requirements.txt).
from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import sys

import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
OUT = os.path.join(REPO_ROOT, "public", "info-publicsafety.json")
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")

# NFIRS incident-type series (first digit of the [NNN] code) -> dashboard label.
# Series 2/8/9 are rare; they fold into "Other".
NFIRS_SERIES = {
    "1": "Fire",
    "2": "Other",                    # overpressure rupture / explosion
    "3": "Rescue & EMS",
    "4": "Hazardous condition",
    "5": "Service call",
    "6": "Good intent",
    "7": "False alarm / false call",
    "8": "Other",                    # severe weather / natural disaster
    "9": "Other",                    # special incident
}
# Display order for the "by type" bar chart (most-relevant civic framing first;
# actual values drive the sort, this is only the tie-break / canonical order).
CATEGORY_ORDER = [
    "Fire", "Rescue & EMS", "Hazardous condition",
    "Service call", "Good intent", "False alarm / false call", "Other",
]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

CODE_RE = re.compile(r"\[(\d{3})\]")          # "[131] Passenger vehicle fire" -> 131
# Address-ish patterns we refuse to publish (PII guard). The street suffix must
# be its own token (a number, then 1-5 words, then e.g. "Rd"/"Ave") so prose like
# "2025 is the first" cannot match on the "st" inside "first".
_STREET = (r"\b\d{1,6}\s+(?:[A-Za-z0-9.\-]+\s+){1,5}"
           r"(?:St|Street|Ave|Avenue|Rd|Road|Dr|Drive|Blvd|Ln|Lane|Ct|Court|"
           r"Way|Hwy|Highway|Pkwy|Pl|Place|Cir|Circle|Ter|Terrace)\b")
PII_PATTERNS = [
    re.compile(_STREET, re.IGNORECASE),
    re.compile(r"\bMichigan\b", re.IGNORECASE),
    re.compile(r"\b48\d{3}\b"),               # local ZIP codes
]


def _sheet_rows(path: str, sheet: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{sheet}' not in {os.path.basename(path)} "
                 f"(found: {wb.sheetnames})")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows


def _header_index(rows: list[tuple]) -> dict:
    return {str(h).strip(): i for i, h in enumerate(rows[0]) if h is not None}


def _cover_year(path: str) -> int | None:
    """Pull the reporting year from a report's Cover sheet ('... 1/1/2025 ...')."""
    for row in _sheet_rows(path, "Cover"):
        for cell in row:
            m = re.search(r"to\s+\d{1,2}/\d{1,2}/(\d{4})", str(cell or ""))
            if m:
                return int(m.group(1))
    return None


def parse_type_report(path: str) -> dict:
    """Aggregate #1390 by NFIRS series. Reads Incident Type only -- NOT Location."""
    rows = _sheet_rows(path, "Source Data")
    idx = _header_index(rows)
    if "Incident Type" not in idx:
        sys.exit(f"ERROR: no 'Incident Type' column in {os.path.basename(path)}")
    it_i = idx["Incident Type"]
    by_category: dict[str, int] = {}
    total_fires = building_fires = total = 0
    for r in rows[1:]:
        raw = r[it_i] if it_i < len(r) else None
        if raw is None:
            continue
        m = CODE_RE.search(str(raw))
        if not m:
            continue
        code = m.group(1)
        label = NFIRS_SERIES.get(code[0], "Other")
        by_category[label] = by_category.get(label, 0) + 1
        total += 1
        if code[0] == "1":
            total_fires += 1
            if code == "111":          # NFIRS 111 = building fire specifically
                building_fires += 1
    return {
        "year": _cover_year(path),
        "by_category": by_category,
        "total": total,
        "total_fires": total_fires,
        "building_fires": building_fires,
    }


def _parse_dt(value) -> datetime.datetime | None:
    if isinstance(value, datetime.datetime):
        return value
    for fmt in ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def parse_year_report(path: str) -> dict:
    """Read #22: per-year totals + the latest full year's monthly spread, and
    flag any partial (mid-year go-live) years from the incident dates."""
    src = _sheet_rows(path, "Source Data")
    idx = _header_index(src)
    dt_i = idx.get("Incident Date/Time")
    if dt_i is None:
        sys.exit(f"ERROR: no 'Incident Date/Time' column in {os.path.basename(path)}")
    per_year: dict[int, int] = {}
    first_month: dict[int, int] = {}
    monthly: dict[int, list[int]] = {}
    for r in src[1:]:
        d = _parse_dt(r[dt_i] if dt_i < len(r) else None)
        if not d:
            continue
        per_year[d.year] = per_year.get(d.year, 0) + 1
        first_month[d.year] = min(first_month.get(d.year, 12), d.month)
        monthly.setdefault(d.year, [0] * 12)[d.month - 1] += 1
    if not per_year:
        sys.exit("ERROR: no parseable incident dates in the year report")
    years = sorted(per_year)
    full_years = [y for y in years if first_month[y] == 1]
    headline = max(full_years) if full_years else max(years)
    partial = [{"year": y, "count": per_year[y], "first_month": first_month[y]}
               for y in years if first_month[y] != 1]
    return {
        "per_year": per_year,
        "headline_year": headline,
        "headline_monthly": monthly[headline],
        "partial_years": partial,
        "years": years,
    }


def build_panel(types: dict, yearly: dict) -> dict:
    year = types["year"] or yearly["headline_year"]
    cats = types["by_category"]

    stats = [
        {"label": "Total responses", "value": f"{types['total']:,}",
         "hint": f"{year} (first full year)"},
        {"label": "Fire responses", "value": f"{types['total_fires']:,}",
         "hint": f"incl. {types['building_fires']} building fires"},
        {"label": "False alarms", "value": f"{cats.get('False alarm / false call', 0):,}",
         "hint": "mostly detector/alarm malfunctions"},
        # Hazardous-condition count is NOT a stat here: the FD-category by-type
        # chart (build_fire_trends.py) shows "Hazardous Conditions" authoritatively
        # from the Chief's workbook, and the NFIRS series-4 count differs slightly,
        # so a duplicate stat card would read as an inconsistency.
    ]

    # The current-year by-TYPE breakdown comes from the Fire Chief's workbooks
    # (FD plain-language categories, with history) via build_fire_trends.py, so
    # the NFIRS series is not charted here -- only the current-year month spread.
    charts = [
        {"type": "trend", "title": f"Responses by month ({year})", "unit": "",
         "points": [{"x": MONTHS[i], "y": yearly["headline_monthly"][i]}
                    for i in range(12)]},
    ]

    notes = [
        "Aggregates from the City of Burton Fire Department records system "
        "(Emergency Networking RMS). Counts only -- no addresses, names, or "
        "individual incidents.",
        "These are fire/rescue incident counts; EMS-coded calls are few in this "
        "dataset (medical transport may be recorded separately).",
    ]
    for p in yearly["partial_years"]:
        notes.append(
            f"{p['year']} is a partial year ({MONTHS[p['first_month'] - 1]}-Dec, "
            f"{p['count']:,} responses) -- the FD adopted this records system mid-"
            f"{p['year']}. {year} is the first full calendar year.")
    return {
        "title": "Burton Fire & Rescue",
        "subtitle": f"Fire & Rescue responses, {year}",
        "logo": "cityofburton_firedeptlogo_nobackground.png",
        "stats": stats,
        "charts": charts,
        "source": f"City of Burton Fire Department incident records "
                  f"(Emergency Networking RMS), calendar year {year}.",
        "links": [],
        "notes": notes,
    }


def assert_no_pii(panel: dict) -> None:
    blob = json.dumps(panel)
    for pat in PII_PATTERNS:
        m = pat.search(blob)
        if m:
            sys.exit(f"PII GUARD TRIPPED: output contains '{m.group(0)}'. "
                     "Refusing to write. An address/location leaked into the panel.")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--type-report",
                    default=os.path.join(DOWNLOADS,
                        "incident-count-and-percentage-per-incident-type.xlsx"),
                    help="#1390 export (.xlsx)")
    ap.add_argument("--year-report",
                    default=os.path.join(DOWNLOADS, "incidents-per-year.xlsx"),
                    help="#22 export (.xlsx)")
    ap.add_argument("--trends-cache",
                    default=os.path.join(os.path.dirname(__file__), "fire-trends.json"),
                    help="multi-year trend fragment from build_fire_trends.py; "
                         "merged in when present so a from-scratch rebuild keeps "
                         "the historical charts. Optional.")
    ap.add_argument("--out", default=OUT, help="output JSON path")
    args = ap.parse_args()

    for label, path in (("type", args.type_report), ("year", args.year_report)):
        if not os.path.isfile(path):
            sys.exit(f"ERROR: {label} report not found: {path}")

    types = parse_type_report(args.type_report)
    yearly = parse_year_report(args.year_report)

    # Cross-check the two reports agree on the headline-year total.
    hy = types["year"]
    if hy and hy in yearly["per_year"] and yearly["per_year"][hy] != types["total"]:
        print(f"WARNING: #1390 total ({types['total']}) != #22 {hy} count "
              f"({yearly['per_year'][hy]}). Using #1390 for categories.",
              file=sys.stderr)

    panel = build_panel(types, yearly)

    # Fold in the Fire Chief's multi-year trend charts when the cache is present
    # (produced by build_fire_trends.py from the OneDrive summary workbooks), so
    # a from-scratch rebuild keeps the historical charts rather than dropping them.
    if args.trends_cache and os.path.isfile(args.trends_cache):
        from build_fire_trends import merge_into_panel
        with open(args.trends_cache, encoding="utf-8") as f:
            panel = merge_into_panel(panel, json.load(f))

    assert_no_pii(panel)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(panel, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {args.out}")
    print(f"  year={panel['subtitle']}")
    print(f"  total={types['total']}  fires={types['total_fires']} "
          f"(building={types['building_fires']})")
    print(f"  categories: " + ", ".join(f"{s['label']}={s['value']}"
                                         for s in panel['charts'][0]['series']))
    print(f"  years in #22: {yearly['per_year']}")


if __name__ == "__main__":
    main()
