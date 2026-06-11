# Fold the Fire Chief's annual summary workbooks into the Public Safety
# (Fire & Rescue) dashboard as MULTI-YEAR trend charts.
#
# Source: five hand-maintained summary workbooks the Fire Chief's office keeps
# (City of Burton OneDrive "ExploreBurton App/"). These are pure AGGREGATES
# (counts by year / station / month) -- no addresses, names, or per-incident
# rows -- so unlike the Emergency Networking exports the dashboard's base panel
# is built from, they carry no PII and may be committed as a derived fragment.
#
# This tool reads them and ADDS historical trend charts to the existing panel:
#   * Annual Stats.xlsx          -> Total calls for service by year (2014-2025)
#   * Station Comparison.xlsx     -> Calls by station area (latest full year)
#   * Monthly Average Worksheet   -> Busiest months (multi-year monthly average)
#   * Annual Comparison.xlsx / Quarterly Statistics.xlsx -> reviewed, NOT charted
#     (Quarterly is redundant with the annual + monthly views; the mutual-aid
#     series from Annual Comparison was removed at the department's request).
#
# GOVERNANCE:
#   * The current-year snapshot (stats + "by type" NFIRS chart) stays as the
#     Emergency Networking base panel already has it. These additions are
#     historical TIME-SERIES only; they never restate the NFIRS taxonomy (the
#     workbook's plain-language categories don't reconcile 1:1 with NFIRS
#     series, so juxtaposing them would invite "these don't add up").
#   * Year totals come from Annual Stats / Station Comparison (2025 = 758,
#     matching the live headline). The Monthly worksheet (NFIRS-station based,
#     2025 sums to 748) is used ONLY for month-to-month SHAPE (seasonality),
#     never for an annual total -- the two sources are never crossed in a chart.
#   * The merge is idempotent (charts/notes this tool owns are replaced, not
#     duplicated) and the same PII guard from build_publicsafety runs on the
#     final panel before it is written.
#
# Re-runnable:
#   python tools/build_fire_trends.py
#   python tools/build_fire_trends.py --src "PATH\to\ExploreBurton App"
#
# Requires: openpyxl (see tools/requirements.txt).
from __future__ import annotations

import argparse
import json
import os
import sys

import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PANEL = os.path.join(REPO_ROOT, "public", "info-publicsafety.json")
CACHE = os.path.join(os.path.dirname(__file__), "fire-trends.json")
# The Fire Chief's workbooks live in the user's OneDrive by default.
DEFAULT_SRC = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - City of Burton", "ExploreBurton App",
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Chart titles and the note marker this tool OWNS. Anything matching is removed
# before re-adding, so repeated runs (or a build_publicsafety merge) never
# duplicate. Titles carry the year range, so they are built, not literals.
HISTORY_MARK = "Multi-year trends are from the Fire Department's annual summary"

# Chart-title prefixes that should NOT linger in an already-merged panel.
# merge_into_panel strips any non-owned chart whose title starts with one of
# these, so a re-run cleans up superseded charts:
#   * "Mutual aid given to neighbors" -- removed at the department's request
#   * "Responses by type ("           -- the NFIRS single-year breakdown,
#                                        replaced by the FD-category "Calls by type"
#   * "Calls by station area ("       -- the single-year bars, replaced by the
#                                        "Calls by station area: YYYY vs YYYY" compare
DEPRECATED_TITLE_PREFIXES = (
    "Mutual aid given to neighbors",
    "Responses by type (",
    "Calls by station area (",
)


# --- workbook readers ------------------------------------------------------

def _rows(path: str) -> list[tuple]:
    """Return all cell rows of the first worksheet (values only)."""
    if not os.path.isfile(path):
        sys.exit(f"ERROR: workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _year_header(row: tuple) -> dict[int, int]:
    """Map {year -> column index} from a header row of 4-digit years."""
    out: dict[int, int] = {}
    for i, cell in enumerate(row):
        if isinstance(cell, (int, float)) and 1900 <= int(cell) <= 2100:
            out[int(cell)] = i
    return out


def _label(cell) -> str:
    return str(cell or "").strip()


def parse_total_by_year(path: str) -> list[dict]:
    """Annual Stats.xlsx -> [{year, total}] for the 'Total Calls for Service'
    row, one point per full calendar year in the header."""
    rows = _rows(path)
    years = _year_header(rows[0])
    if not years:
        sys.exit(f"ERROR: no year header in {os.path.basename(path)}")
    total_row = next(
        (r for r in rows[1:]
         if _label(r[0]).lower().startswith("total calls for service")),
        None,
    )
    if total_row is None:
        sys.exit(f"ERROR: no 'Total Calls for Service' row in {os.path.basename(path)}")
    points = []
    for yr in sorted(years):
        ci = years[yr]
        v = total_row[ci] if ci < len(total_row) else None
        if isinstance(v, (int, float)):
            points.append({"year": yr, "total": int(v)})
    return points


# The Fire Chief's workbooks use the department's own plain-language incident
# categories (not the NFIRS series). Map each row label to a clean display name,
# in operational order; unrecognized rows (e.g. "Total Calls for Service") skip.
CATEGORY_DISPLAY = [
    ("structure fire", "Structure Fires"),
    ("vehicle fire", "Vehicle Fires"),
    ("dumpster", "Brush & Trash Fires"),
    ("other fire", "Other Fires"),
    ("extrication", "Rescues & Extrications"),
    ("alarm", "Fire & CO Alarms"),
    ("hazardous", "Hazardous Conditions"),
    ("good intent", "Good Intent / Service"),
    ("open burning", "Open Burning"),
    ("cancelled en route", "Cancelled En Route"),
]


def _display_category(raw: str) -> str | None:
    low = raw.lower()
    for sub, disp in CATEGORY_DISPLAY:
        if sub in low:
            return disp
    return None


def parse_category_by_year(path: str) -> dict[str, dict[int, int]]:
    """Annual Stats.xlsx -> {display_category: {year: count}} for each incident
    category row (the Total row and unrecognized labels are skipped)."""
    rows = _rows(path)
    years = _year_header(rows[0])
    out: dict[str, dict[int, int]] = {}
    for r in rows[1:]:
        disp = _display_category(_label(r[0]))
        if disp is None:
            continue
        by_year = {yr: int(r[ci]) for yr, ci in years.items()
                   if ci < len(r) and isinstance(r[ci], (int, float))}
        if by_year:
            out[disp] = by_year
    return out


def parse_station_by_year(path: str) -> dict[int, list[tuple[str, int]]]:
    """Station Comparison.xlsx -> {year: [(station_label, calls), ...]}."""
    rows = _rows(path)
    header = rows[0]
    year_i = next((i for i, c in enumerate(header)
                   if _label(c).upper() == "YEAR"), None)
    station_cols = [(_label(header[i]), i) for i in range(len(header))
                    if _label(header[i]).upper().startswith("STA")]
    if year_i is None or not station_cols:
        sys.exit(f"ERROR: missing YEAR/STA columns in {os.path.basename(path)}")
    out: dict[int, list[tuple[str, int]]] = {}
    for r in rows[1:]:
        yv = r[year_i] if year_i < len(r) else None
        if not isinstance(yv, (int, float)):
            continue  # skips Total:/Annual Avg: footer rows
        stations = []
        for name, ci in station_cols:
            v = r[ci] if ci < len(r) else None
            if isinstance(v, (int, float)):
                # "STA. #1" -> "Station 1 area"
                label = name.replace("STA.", "Station").replace("#", "").strip()
                stations.append((f"{label} area", int(v)))
        if stations:
            out[int(yv)] = stations
    return out


def _baseline_year(years: list[int], latest: int, span: int = 10) -> int:
    """The year `span` before `latest` if present, else the earliest year --
    used as the 'then' column of a then-vs-now comparison."""
    return latest - span if (latest - span) in years else min(years)


def parse_monthly_average(path: str) -> dict:
    """Monthly Average Worksheet.xlsx -> {'years': [..], 'avg': [12 floats]}.
    Averages only FULL calendar years (all 12 months present), excluding any
    partial current year, so the seasonality shape is clean and self-consistent."""
    rows = _rows(path)
    # header: ['', January, February, ... December, Total] -> month columns 1..12
    month_cols = list(range(1, 13))
    sums = [0.0] * 12
    counts = [0] * 12
    used_years: list[int] = []
    for r in rows[1:]:
        yv = r[0] if r else None
        if not isinstance(yv, (int, float)):
            continue  # skips AVG. / blank / footnote rows
        vals = [r[c] if c < len(r) else None for c in month_cols]
        if not all(isinstance(v, (int, float)) for v in vals):
            continue  # partial year (e.g. current year through March) -> skip
        used_years.append(int(yv))
        for m in range(12):
            sums[m] += float(vals[m])
            counts[m] += 1
    avg = [round(sums[m] / counts[m], 1) if counts[m] else 0.0 for m in range(12)]
    return {"years": sorted(used_years), "avg": avg}


# --- chart assembly --------------------------------------------------------

def _compare_rows(latest: int, base: int,
                  series: list[tuple[str, dict[int, int]]]) -> list[dict]:
    """Build CompareRow dicts (latest emphasized first, then the baseline year)
    for a then-vs-now grouped-bar chart."""
    return [
        {"label": label, "unit": "",
         "values": [{"name": str(latest), "value": by_year.get(latest, 0)},
                    {"name": str(base), "value": by_year.get(base, 0)}]}
        for label, by_year in series
    ]


def build_fragment(src: str) -> dict:
    """Read the workbooks under `src` and return {charts, stats, notes} to merge."""
    totals = parse_total_by_year(os.path.join(src, "Annual Stats.xlsx"))
    categories = parse_category_by_year(os.path.join(src, "Annual Stats.xlsx"))
    stations = parse_station_by_year(os.path.join(src, "Station Comparison.xlsx"))
    monthly = parse_monthly_average(os.path.join(src, "Monthly Average Worksheet.xlsx"))

    latest = totals[-1]["year"] if totals else max(stations)
    charts: list[dict] = []

    # Calls by type -- the FD's own categories, latest vs ~a decade ago (the only
    # multi-year by-type data; NFIRS series are not carried over time). Biggest
    # current category first.
    if categories:
        cat_years = sorted({y for by in categories.values() for y in by})
        base = _baseline_year(cat_years, latest)
        ordered = sorted(categories.items(), key=lambda kv: -kv[1].get(latest, 0))
        charts.append({
            "type": "compare",
            "title": f"Calls by type: {latest} vs {base}",
            "rows": _compare_rows(latest, base, ordered),
        })

    if totals:
        y0, y1 = totals[0]["year"], totals[-1]["year"]
        chart = {
            "type": "trend",
            "title": f"Total calls for service by year ({y0}-{y1})",
            "unit": "",
            "points": [{"x": str(p["year"]), "y": p["total"]} for p in totals],
        }
        # Mark the pandemic year on the yearly series (only if 2020 is a real point).
        if any(p["year"] == 2020 for p in totals):
            chart["markers"] = [{"x": "2020", "label": "COVID-19"}]
        charts.append(chart)

    # Calls by station area -- latest vs ~a decade ago (was a single-year bars).
    if stations and latest in stations:
        st_years = sorted(stations)
        base = _baseline_year(st_years, latest)
        base_map = dict(stations.get(base, []))
        series = [(name, {latest: val, base: base_map.get(name, 0)})
                  for name, val in stations[latest]]
        charts.append({
            "type": "compare",
            "title": f"Calls by station area: {latest} vs {base}",
            "rows": _compare_rows(latest, base, series),
        })

    if monthly["years"]:
        ys = monthly["years"]
        span = f"{ys[0]}-{ys[-1]}" if len(ys) > 1 else str(ys[0])
        charts.append({
            "type": "bars",
            "title": f"Busiest months ({span} average)",
            "unit": "",
            "series": [{"label": MONTHS[m], "value": monthly["avg"][m]}
                       for m in range(12)],
        })

    # A trend stat: growth over the available annual span.
    stats: list[dict] = []
    if len(totals) >= 2 and totals[0]["total"] > 0:
        first, last = totals[0], totals[-1]
        pct = round((last["total"] - first["total"]) / first["total"] * 100)
        stats.append({
            "label": f"{last['year'] - first['year']}-year change",
            "value": f"{pct:+d}%",
            "hint": f"{first['total']:,} ({first['year']}) -> "
                    f"{last['total']:,} ({last['year']})",
        })

    note = (
        "Multi-year trends are from the Fire Department's annual summary "
        "workbooks (calls for service by year, type, station, and month), "
        "maintained by the Fire Chief's office. Year totals are full calendar "
        "years; by-type and by-station comparisons show the latest year against "
        "about a decade earlier. "
        "Month averages show seasonality from a separate monthly tally, so "
        "their yearly sum differs slightly from the annual totals."
    )
    return {"charts": charts, "stats": stats, "notes": [note]}


def merge_into_panel(panel: dict, fragment: dict) -> dict:
    """Idempotently fold the fragment's charts/stats/notes into a panel. Charts
    and notes this tool owns are removed first, so repeated runs do not stack."""
    owned_titles = {c["title"] for c in fragment["charts"]}

    def _keep(chart: dict) -> bool:
        t = chart.get("title", "")
        if t in owned_titles:
            return False  # this tool re-adds it below
        return not t.startswith(DEPRECATED_TITLE_PREFIXES)

    panel["charts"] = [c for c in panel.get("charts", []) if _keep(c)] + fragment["charts"]

    owned_stat_labels = {s["label"] for s in fragment["stats"]}
    panel["stats"] = [s for s in panel.get("stats", [])
                      if s.get("label") not in owned_stat_labels] + fragment["stats"]

    notes = [n for n in panel.get("notes", []) if not n.startswith(HISTORY_MARK)]
    # keep the trailing "Draft -- ..." note last if present
    draft = [n for n in notes if n.startswith("Draft")]
    notes = [n for n in notes if not n.startswith("Draft")]
    panel["notes"] = notes + fragment["notes"] + draft
    return panel


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", default=DEFAULT_SRC,
                    help="folder holding the Fire Chief's summary .xlsx workbooks")
    ap.add_argument("--panel", default=PANEL, help="info-publicsafety.json to update")
    ap.add_argument("--cache", default=CACHE,
                    help="committable fragment cache (read by build_publicsafety)")
    args = ap.parse_args()

    fragment = build_fragment(args.src)

    with open(args.cache, "w", encoding="utf-8") as f:
        json.dump(fragment, f, indent=2, ensure_ascii=False)
        f.write("\n")

    if not os.path.isfile(args.panel):
        sys.exit(f"ERROR: base panel not found: {args.panel}. "
                 "Run build_publicsafety.py first.")
    with open(args.panel, encoding="utf-8") as f:
        panel = json.load(f)

    panel = merge_into_panel(panel, fragment)

    # Reuse the same PII guard the base tool enforces.
    from build_publicsafety import assert_no_pii
    assert_no_pii(panel)

    with open(args.panel, "w", encoding="utf-8") as f:
        json.dump(panel, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {args.cache}")
    print(f"Updated {args.panel}")
    print(f"  added {len(fragment['charts'])} charts, "
          f"{len(fragment['stats'])} stat(s)")
    for c in fragment["charts"]:
        n = len(c.get("points") or c.get("series") or c.get("rows") or [])
        print(f"    - {c['title']}  ({n} rows)")


if __name__ == "__main__":
    main()
