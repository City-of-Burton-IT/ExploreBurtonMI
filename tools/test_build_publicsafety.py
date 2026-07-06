# Tests for build_publicsafety.py.
#
# PRIORITY: assert_no_pii is the only guard between raw fire-department incident
# exports (which carry street addresses in a "Location" column this script never
# reads) and the committed public JSON. These tests prove it actually trips on
# address-like strings and other PII patterns, passes clean panel data, and
# document what it does NOT catch (it is a regex guard, not a PII scanner).
#
# The xlsx parsers (parse_type_report / parse_year_report / _cover_year /
# _parse_dt) are exercised against tiny synthetic workbooks built with openpyxl
# so no real RMS export is ever touched. Run: python -m pytest tools/test_build_publicsafety.py -q
import datetime
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(__file__))
import build_publicsafety as bps  # noqa: E402


# --- fixtures: tiny synthetic .xlsx workbooks ----------------------------------

def _make_type_report(tmp_path, incident_types, cover_text="Report 1/1/2025 to 12/31/2025"):
    path = tmp_path / "type_report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Data"
    ws.append(["Incident Type"])
    for t in incident_types:
        ws.append([t])
    cover = wb.create_sheet("Cover")
    cover.append([cover_text])
    wb.save(path)
    return str(path)


def _make_year_report(tmp_path, dates):
    path = tmp_path / "year_report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Data"
    ws.append(["Incident Date/Time"])
    for d in dates:
        ws.append([d])
    wb.save(path)
    return str(path)


# --- assert_no_pii: the guard ---------------------------------------------------

def test_assert_no_pii_passes_clean_panel():
    panel = {
        "title": "Burton Fire & Rescue",
        "stats": [{"label": "Total responses", "value": "758", "hint": "2025 (first full year)"}],
        "notes": ["Counts only -- no addresses, names, or individual incidents."],
    }
    bps.assert_no_pii(panel)  # must not raise


def test_assert_no_pii_trips_on_street_address():
    panel = {"notes": ["Crews responded near 1234 Elm Street last week."]}
    with pytest.raises(SystemExit, match="PII GUARD TRIPPED"):
        bps.assert_no_pii(panel)


def test_assert_no_pii_trips_on_directional_street_address():
    panel = {"notes": ["A fire was reported at 500 S Center Rd."]}
    with pytest.raises(SystemExit, match="PII GUARD TRIPPED"):
        bps.assert_no_pii(panel)


def test_assert_no_pii_trips_on_state_name():
    panel = {"source": "City of Burton, Michigan fire records."}
    with pytest.raises(SystemExit, match="PII GUARD TRIPPED"):
        bps.assert_no_pii(panel)


def test_assert_no_pii_trips_on_local_zip():
    panel = {"notes": ["Coverage area centered on zip 48509."]}
    with pytest.raises(SystemExit, match="PII GUARD TRIPPED"):
        bps.assert_no_pii(panel)


def test_assert_no_pii_does_not_false_positive_on_first_full_year_phrase():
    # Regression check for the false-positive the source comment calls out:
    # "2025 is the first" must not match the street-suffix pattern via "first"'s "st".
    panel = {"notes": ["2025 is the first full year of data under this records system."]}
    bps.assert_no_pii(panel)  # must not raise


def test_assert_no_pii_gap_misses_bare_house_number_without_suffix():
    # Documents a real gap: the street regex requires a recognized suffix word
    # (St/Ave/Rd/...), so a house number with no suffix slips through uncaught.
    panel = {"notes": ["Reported near 1234 Main, no cross street given."]}
    bps.assert_no_pii(panel)  # does NOT raise -- gap, not a guarantee


def test_assert_no_pii_gap_misses_names_and_out_of_state_zips():
    # Documents a real gap: there is no name-detection pattern at all, and the
    # ZIP pattern only matches Burton-area 48xxx codes, so a name or an
    # out-of-state ZIP (with no street suffix alongside it) slips through.
    panel = {"notes": ["Contact John Smith, formerly of Reno, NV 89501."]}
    bps.assert_no_pii(panel)  # does NOT raise -- name and non-48xxx ZIP both miss


def test_build_panel_output_passes_the_pii_guard():
    types = {"year": 2025, "by_category": {"Fire": 63, "Rescue & EMS": 10,
                                            "False alarm / false call": 82},
             "total": 155, "total_fires": 63, "building_fires": 44}
    yearly = {"headline_year": 2025, "headline_monthly": [10] * 12,
              "partial_years": [], "years": [2025], "per_year": {2025: 155}}
    panel = bps.build_panel(types, yearly)
    bps.assert_no_pii(panel)  # a real panel must clear the guard


# --- parse_type_report -----------------------------------------------------------

def test_parse_type_report_aggregates_by_nfirs_series(tmp_path):
    path = _make_type_report(tmp_path, [
        "[111] Building fire",
        "[111] Building fire",
        "[131] Passenger vehicle fire",
        "[321] EMS call",
        "[743] Alarm system sounded",
        "no code here",
        None,
    ])
    out = bps.parse_type_report(path)
    assert out["total"] == 5  # "no code here" and None are skipped
    assert out["total_fires"] == 3  # 111, 111, 131 -> series "1"
    assert out["building_fires"] == 2  # exact code 111
    assert out["by_category"] == {
        "Fire": 3, "Rescue & EMS": 1, "False alarm / false call": 1,
    }
    assert out["year"] == 2025  # from the Cover sheet


def test_parse_type_report_missing_column_exits(tmp_path):
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Data"
    ws.append(["Not Incident Type"])
    ws.append(["[111] Building fire"])
    wb.save(path)
    with pytest.raises(SystemExit, match="no 'Incident Type' column"):
        bps.parse_type_report(str(path))


def test_cover_year_no_match_returns_none(tmp_path):
    path = _make_type_report(tmp_path, ["[111] Building fire"], cover_text="no date pattern here")
    assert bps._cover_year(path) is None


# --- parse_year_report ------------------------------------------------------------

def test_parse_year_report_full_and_partial_years(tmp_path):
    dates = [
        datetime.datetime(2023, 6, 15),  # partial year: first_month=6
        datetime.datetime(2023, 7, 1),
        datetime.datetime(2024, 1, 5),
        datetime.datetime(2024, 3, 20),
        datetime.datetime(2025, 1, 10),
    ]
    path = _make_year_report(tmp_path, dates)
    out = bps.parse_year_report(path)
    assert out["per_year"] == {2023: 2, 2024: 2, 2025: 1}
    assert out["years"] == [2023, 2024, 2025]
    assert out["headline_year"] == 2025  # latest full (Jan-starting) year
    assert out["partial_years"] == [{"year": 2023, "count": 2, "first_month": 6}]
    assert len(out["headline_monthly"]) == 12
    assert out["headline_monthly"][0] == 1  # the one Jan-2025 row
    assert sum(out["headline_monthly"]) == 1


def test_parse_year_report_falls_back_to_latest_year_when_none_full(tmp_path):
    dates = [datetime.datetime(2023, 6, 1), datetime.datetime(2024, 3, 1)]
    path = _make_year_report(tmp_path, dates)
    out = bps.parse_year_report(path)
    assert out["headline_year"] == 2024  # no full years -> just the latest
    assert {p["year"] for p in out["partial_years"]} == {2023, 2024}


def test_parse_year_report_string_dates_are_parsed(tmp_path):
    path = _make_year_report(tmp_path, ["1/5/2025 08:30:00", "1/6/2025 09:00"])
    out = bps.parse_year_report(path)
    assert out["per_year"] == {2025: 2}


def test_parse_year_report_missing_column_exits(tmp_path):
    path = tmp_path / "bad_year.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Data"
    ws.append(["Not A Date Column"])
    wb.save(path)
    with pytest.raises(SystemExit, match="no 'Incident Date/Time' column"):
        bps.parse_year_report(str(path))


def test_parse_year_report_no_parseable_dates_exits(tmp_path):
    path = _make_year_report(tmp_path, ["not a date", None, "also not a date"])
    with pytest.raises(SystemExit, match="no parseable incident dates"):
        bps.parse_year_report(path)


# --- _parse_dt / _header_index --------------------------------------------------

@pytest.mark.parametrize("value", [
    "6/15/2023 10:30:00",
    "6/15/2023 10:30",
    "6/15/2023",
])
def test_parse_dt_accepts_known_formats(value):
    d = bps._parse_dt(value)
    assert d is not None
    assert (d.year, d.month, d.day) == (2023, 6, 15)


def test_parse_dt_passthrough_datetime_object():
    now = datetime.datetime(2024, 1, 1, 12, 0)
    assert bps._parse_dt(now) is now


def test_parse_dt_rejects_garbage():
    assert bps._parse_dt("not a date") is None
    assert bps._parse_dt(None) is None


def test_header_index_maps_names_to_column_positions():
    rows = [("Incident Type", None, "Incident Date/Time")]
    assert bps._header_index(rows) == {"Incident Type": 0, "Incident Date/Time": 2}


# --- build_panel -------------------------------------------------------------------

def test_build_panel_stats_and_partial_year_note():
    types = {"year": 2025, "by_category": {"Fire": 63, "False alarm / false call": 82},
             "total": 145, "total_fires": 63, "building_fires": 44}
    yearly = {"headline_year": 2025, "headline_monthly": list(range(12)),
              "partial_years": [{"year": 2020, "count": 300, "first_month": 6}],
              "years": [2020, 2025], "per_year": {2020: 300, 2025: 145}}
    panel = bps.build_panel(types, yearly)
    stats = {s["label"]: s["value"] for s in panel["stats"]}
    assert stats["Total responses"] == "145"
    assert stats["Fire responses"] == "63"
    assert stats["False alarms"] == "82"
    assert panel["charts"][0]["points"][0] == {"x": "Jan", "y": 0}
    assert any("2020 is a partial year" in n for n in panel["notes"])
    assert panel["subtitle"] == "Fire & Rescue responses, 2025"


def test_build_panel_falls_back_to_yearly_headline_when_types_year_missing():
    types = {"year": None, "by_category": {}, "total": 0, "total_fires": 0, "building_fires": 0}
    yearly = {"headline_year": 2024, "headline_monthly": [0] * 12,
              "partial_years": [], "years": [2024], "per_year": {2024: 0}}
    panel = bps.build_panel(types, yearly)
    assert panel["subtitle"] == "Fire & Rescue responses, 2024"
