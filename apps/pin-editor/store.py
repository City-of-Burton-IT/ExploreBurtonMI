# Pin-editor IO + pipeline + git layer. Resolves repo paths, loads/saves the source
# and generated GeoJSON (preserving each file's existing line endings so git diffs stay
# minimal), runs the data pipeline, and does the git publish. Kept separate from the
# pure routing in edits.py so the routing stays trivially testable.
from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import excel

# apps/pin-editor/store.py -> repo root is two levels up.
REPO_ROOT = Path(__file__).resolve().parents[2]
FACILITIES = REPO_ROOT / "pipeline" / "data" / "facilities.geojson"
OVERRIDES = REPO_ROOT / "pipeline" / "data" / "overrides.json"
DATA = REPO_ROOT / "public" / "data.geojson"
BOUNDARY = REPO_ROOT / "public" / "boundary.geojson"
PIPELINE_DIR = REPO_ROOT / "pipeline"
_PIPELINE_PY = PIPELINE_DIR / ".venv" / "Scripts" / "python.exe"

# Files the publish step stages (source + everything the pipeline regenerates).
PUBLISH_PATHS = [FACILITIES, OVERRIDES, DATA, BOUNDARY]

# City bounding box (min_lng, min_lat, max_lng, max_lat) -- mirrors the pipeline's
# validation extent; used to warn on out-of-city adds/moves.
CITY_BBOX = (-83.85, 42.85, -83.40, 43.15)

# Controlled category vocabulary (the distinct categories the public map uses).
CATEGORIES = [
    "Government", "Public Safety", "Public Works", "Parks & Recreation",
    "Community Services", "Education", "Faith", "Health & Medical",
    "Dining", "Grocery & Food", "Retail & Shopping", "Automotive",
    "Professional & Personal Services", "Financial", "Arts & Entertainment",
    "Lodging",
]


def _read_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _compact_geometry(text: str) -> str:
    """Collapse each pretty-printed `"geometry": { ... }` block onto one line to match
    facilities.geojson's hand-maintained style -- so unchanged features produce no diff.
    A Point geometry has no nested braces, so `\\{[^}]*\\}` matches it cleanly."""
    def repl(m: "re.Match") -> str:
        obj = json.loads(m.group(1))
        compact = json.dumps(obj, separators=(", ", ": "))
        return '"geometry": { ' + compact[1:-1] + " }"
    return re.sub(r'"geometry":\s*(\{[^}]*\})', repl, text)


def _write_json(path: Path, obj: dict, compact_geometry: bool = False) -> None:
    """Write pretty JSON, preserving the file's existing newline style (LF vs CRLF).
    With compact_geometry, geometry objects are kept inline (facilities.geojson style)."""
    raw = path.read_bytes() if path.exists() else b""
    eol = "\r\n" if b"\r\n" in raw else "\n"
    text = json.dumps(obj, indent=2, ensure_ascii=False)
    if compact_geometry:
        text = _compact_geometry(text)
    if not text.endswith("\n"):
        text += "\n"
    if eol == "\r\n":
        text = text.replace("\n", "\r\n")
    # newline="" so Python does not translate the line endings we just set.
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(text)


def load_facilities() -> dict:
    return _read_json(FACILITIES)


def load_overrides() -> dict:
    return _read_json(OVERRIDES)


def load_data() -> dict:
    return _read_json(DATA)


def load_boundary() -> dict:
    return _read_json(BOUNDARY)


def save_sources(facilities: dict, overrides: dict) -> None:
    _write_json(FACILITIES, facilities, compact_geometry=True)
    _write_json(OVERRIDES, overrides)


def build_pins_workbook(rows: list) -> BytesIO:
    """Build the editable pins .xlsx: a Pins sheet (frozen bold header, shaded read-only
    id/source columns, a category dropdown and a delete yes/no dropdown) plus a hidden
    Lists sheet backing the category dropdown. Returns a BytesIO ready to send."""
    cols = excel.COLUMNS
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Pins"

    lists = wb.create_sheet("Lists")
    for r, cat in enumerate(CATEGORIES, start=1):
        lists.cell(row=r, column=1, value=cat)
    lists.sheet_state = "hidden"

    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    def _cell(v):  # openpyxl only accepts scalars; stringify anything else
        return v if v is None or isinstance(v, (str, int, float)) else str(v)
    for row in rows:
        ws.append([_cell(row.get(c, "")) for c in cols])

    last = len(rows) + 1 + 300                         # spare rows for new pins
    cat_col = get_column_letter(cols.index("category") + 1)
    del_col = get_column_letter(cols.index("delete") + 1)
    dv_cat = DataValidation(type="list", formula1=f"Lists!$A$1:$A${len(CATEGORIES)}", allow_blank=True)
    dv_del = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_del)
    dv_cat.add(f"{cat_col}2:{cat_col}{last}")
    dv_del.add(f"{del_col}2:{del_col}{last}")

    shade = PatternFill("solid", fgColor="EEEEEE")
    for name in ("id", "source"):
        ci = cols.index(name) + 1
        for r in range(1, len(rows) + 2):
            ws.cell(row=r, column=ci).fill = shade

    widths = {"id": 24, "source": 11, "name": 28, "category": 26, "address": 32,
              "phone": 16, "website": 32, "hours": 22, "lat": 11, "lng": 11, "delete": 8}
    for ci, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(c, 14)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def read_pins_rows(fileobj) -> list:
    """Read an uploaded pins .xlsx into a list of row dicts keyed by header name."""
    wb = openpyxl.load_workbook(fileobj, data_only=True, read_only=True)
    ws = wb["Pins"] if "Pins" in wb.sheetnames else wb.active
    assert ws is not None
    rows, header = [], None
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = ["" if h is None else str(h).strip() for h in r]
            continue
        d = {h: (r[i] if i < len(r) else None) for i, h in enumerate(header) if h}
        rows.append(d)
    return rows


def _pipeline_python() -> str:
    return str(_PIPELINE_PY) if _PIPELINE_PY.exists() else sys.executable


def run_pipeline() -> dict:
    """Run pipeline/run.py (cached sources, no network). Returns ok/stdout/stderr/counts."""
    before = len(load_data().get("features", [])) if DATA.exists() else 0
    proc = subprocess.run(
        [_pipeline_python(), "run.py"],
        cwd=str(PIPELINE_DIR), capture_output=True, text=True,
    )
    after = len(load_data().get("features", [])) if DATA.exists() else 0
    return {
        "ok": proc.returncode == 0,
        "returncode": proc.returncode,
        "stdout": proc.stdout[-4000:],
        "stderr": proc.stderr[-4000:],
        "before": before,
        "after": after,
    }


def _git(*args: str) -> subprocess.CompletedProcess:
    return subprocess.run(
        ["git", "-C", str(REPO_ROOT), *args], capture_output=True, text=True,
    )


def git_diff_stat() -> str:
    rels = [str(p.relative_to(REPO_ROOT)).replace("\\", "/") for p in PUBLISH_PATHS]
    proc = _git("diff", "--stat", "--", *rels)
    return proc.stdout.strip() or "(no changes)"


_SSN_PATTERN = re.compile(r"(?<!\d)\d{3}-\d{2}-\d{4}(?!\d)")
_FORBIDDEN_PUBLIC_FIELD_PATTERN = re.compile(
    r'^\+\s*"(?:account(?:number)?|dateofbirth|dob|email|owner(?:name)?|'
    r'parcel(?:id|number)?|ssn|taxpayer)"\s*:',
    re.IGNORECASE,
)


def _pii_findings(staged_diff: str) -> list[str]:
    """Return finding labels for sensitive values added to public data files.

    Finding labels intentionally omit the matching content so browser responses and
    logs never repeat a sensitive value.
    """
    findings: list[str] = []
    for line in staged_diff.splitlines():
        if not line.startswith("+") or line.startswith("+++"):
            continue
        if _SSN_PATTERN.search(line) and "SSN-like value" not in findings:
            findings.append("SSN-like value")
        if (
            _FORBIDDEN_PUBLIC_FIELD_PATTERN.search(line)
            and "forbidden personal/property field" not in findings
        ):
            findings.append("forbidden personal/property field")
    return findings


def _scan_staged_for_secrets() -> dict:
    """Fail closed unless secret and PII checks approve the staged public diff."""
    executable = shutil.which("gitleaks")
    if not executable:
        return {
            "ok": False,
            "detail": "gitleaks is required before publishing; install it and try again.",
        }
    proc = subprocess.run(
        [executable, "git", str(REPO_ROOT), "--staged", "--no-banner"],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        return {
            "ok": False,
            "detail": "Secret preflight failed. Run gitleaks on the staged diff for details.",
        }
    rels = [str(path.relative_to(REPO_ROOT)).replace("\\", "/") for path in PUBLISH_PATHS]
    diff = _git("diff", "--cached", "--unified=0", "--no-color", "--", *rels)
    if diff.returncode != 0:
        return {
            "ok": False,
            "detail": "PII preflight could not inspect the staged public-data diff.",
        }
    findings = _pii_findings(diff.stdout)
    if findings:
        return {
            "ok": False,
            "detail": "PII preflight blocked: " + ", ".join(findings) + ".",
        }
    return {"ok": True, "detail": "no secrets or PII found"}


def git_publish(message: str) -> dict:
    """Publish only allowlisted data files after index and secret preflights."""
    rels = [str(p.relative_to(REPO_ROOT)).replace("\\", "/") for p in PUBLISH_PATHS]
    staged = _git("diff", "--cached", "--name-only", "-z", "--")
    if staged.returncode != 0:
        return {"ok": False, "step": "preflight", "detail": staged.stderr.strip()}
    staged_paths = {path for path in staged.stdout.split("\0") if path}
    unrelated = sorted(staged_paths - set(rels))
    if unrelated:
        return {
            "ok": False,
            "step": "preflight",
            "detail": "Refusing to publish with unrelated staged files: " + ", ".join(unrelated),
        }
    add = _git("add", "--", *rels)
    if add.returncode != 0:
        return {"ok": False, "step": "add", "detail": add.stderr.strip()}
    scan = _scan_staged_for_secrets()
    if not scan["ok"]:
        return {"ok": False, "step": "scan", "detail": scan["detail"]}
    commit = _git("commit", "--only", "-m", message, "--", *rels)
    if commit.returncode != 0:
        out = (commit.stdout + commit.stderr).lower()
        if "nothing to commit" in out:
            return {"ok": False, "step": "commit", "detail": "Nothing to commit (no staged changes)."}
        return {"ok": False, "step": "commit", "detail": (commit.stdout + commit.stderr).strip()}
    push = _git("push")
    if push.returncode != 0:
        return {"ok": False, "step": "push", "detail": (push.stdout + push.stderr).strip(),
                "committed": True}
    return {"ok": True, "detail": (commit.stdout + push.stderr).strip()}
